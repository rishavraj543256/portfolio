import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import shutil
import os
from datetime import datetime
from openpyxl.utils import column_index_from_string
from openpyxl.cell.cell import MergedCell
from sum_logic import sum_column_after_marker

# Label text -> (data key, target column letter)
labels_with_column = {
    "Report Number":        ("report_number", "D", "D:G"),
    "Zone":                 ("zone", "D", "D:G"),
    "Unit":                 ("unit", "D", "D:G"),
    "Customer Code of the Distributor": ("customer_code", "D", "D:G"),
    "Name of the Distributor": ("distributor_name", "D", "D:G"),
    "Address 1":            ("address1", "D", "D:G"),
    "Address 2":            ("address2", "L", "L:O"),
    "City, State":          ("city_state", "L", "L:O"),
    "Audit Firm":           ("audit_firm", "L", "L:O"),
    "Audit Team Lead :":      ("audit_team_lead", "L", "L:O"),
    "Contact No":           ("contact_no", "L", "L:O"),
    "Date of audit":        ("date_of_audit", "L", "L:O")
}


def copy_template():
    template_path = os.path.join("template", "template.xlsx")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    new_filename = f"template_copy_{timestamp}.xlsx"
    
    if not os.path.exists(template_path):
        print(f"Template file not found at: {template_path}")
        return None
    
    try:
        shutil.copy2(template_path, new_filename)
        print(f"Excel file copied successfully to: {new_filename}")
        return new_filename
    except Exception as e:
        print(f"Error copying file: {e}")
        return None

def find_and_fill(ws, label, column_letter, value, merge_range=None):
    """Find row by label and fill value in fixed column and optionally merge."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == label.lower():
                row_number = cell.row

                if merge_range:
                    start_col, end_col = merge_range.split(":")
                    start_col_idx = column_index_from_string(start_col)
                    end_col_idx = column_index_from_string(end_col)
                    cell_range = f"{start_col}{row_number}:{end_col}{row_number}"

                    # First unmerge if already merged
                    if cell_range in ws.merged_cells.ranges:
                        ws.unmerge_cells(cell_range)

                    ws.merge_cells(cell_range)
                    ws.cell(row=row_number, column=start_col_idx).value = value
                    ws.cell(row=row_number, column=start_col_idx).alignment = Alignment(vertical='center', horizontal='left')
                else:
                    ws[f"{column_letter}{row_number}"].value = value

                return


def fill_header_section(ws, matched_row, serial_number):
    """Fill headers dynamically based on label names"""
    # Prepare values as per required keys
    customer_name = str(matched_row['Customer Name'])
    if '-' in customer_name:
        number_part, string_part = customer_name.split('-', 1)
    else:
        number_part = string_part = customer_name

    distributor_data = {
        "report_number": serial_number,
        "zone": matched_row['Zone'],
        "unit": "",
        "customer_code": number_part.strip(),
        "distributor_name": string_part.strip(),
        "address1": matched_row['Address 1'],
        "address2": matched_row['Address 2'],
        "city_state": f"{matched_row['District']}, {matched_row['State Name']}",
        "audit_firm": "RUTUL SHAH & ASSOCIATES",
        "audit_team_lead": "",
        "contact_no": "",
        "date_of_audit": ""
    }

    for label, (key, column, merge_range) in labels_with_column.items():
        find_and_fill(ws, label, column, distributor_data.get(key, ""), merge_range)


def process_single_report(ws, df1_path, df2_path, start_row, serial_number):
    """Process a single report beginning at the given start row"""
    try:
        df1 = pd.read_excel(df1_path)
        df2 = pd.read_excel(df2_path)
        
        for code in df1['Distributor code']:
            match = df2[df2['Codes'] == code]
            if not match.empty:
                matched_row = match.iloc[0]
                
                # Copy template section headers to the current position
                # This assumes template headers are always in the same position
                if start_row > 1:  # Not for the first report which uses the original template
                    # Copy the header rows (assuming they're in rows 1-14)
                    for i in range(1, 15):
                        for j in range(1, 20):  # Assume columns A through T
                            source_cell = ws.cell(row=i, column=j)
                            target_cell = ws.cell(row=start_row + i - 1, column=j)
                            
                            # Copy value and formatting
                            target_cell.value = source_cell.value
                            
                            # Copy style attributes if they exist
                            if source_cell.has_style:
                                target_cell.font = source_cell.font
                                target_cell.border = source_cell.border
                                target_cell.fill = source_cell.fill
                                target_cell.number_format = source_cell.number_format
                                target_cell.protection = source_cell.protection
                                target_cell.alignment = source_cell.alignment
                
                # Adjust for the new position
                header_offset = start_row - 1 if start_row > 1 else 0
                
                # Prepare values as per required keys
                customer_name = str(matched_row['Customer Name'])
                if '-' in customer_name:
                    number_part, string_part = customer_name.split('-', 1)
                else:
                    number_part = string_part = customer_name

                distributor_data = {
                    "report_number": serial_number,
                    "zone": matched_row['Zone'],
                    "unit": "",
                    "customer_code": number_part.strip(),
                    "distributor_name": string_part.strip(),
                    "address1": matched_row['Address 1'],
                    "address2": matched_row['Address 2'],
                    "city_state": f"{matched_row['District']}, {matched_row['State Name']}",
                    "audit_firm": "RUTUL SHAH & ASSOCIATES",
                    "audit_team_lead": "",
                    "contact_no": "",
                    "date_of_audit": ""
                }

                # Fill header section for this report
                for label, (key, column, merge_range) in labels_with_column.items():
                    if start_row == 1:  # First report - use original template positions
                        find_and_fill(ws, label, column, distributor_data.get(key, ""), merge_range)
                    else:
                        # For subsequent reports, we need to adjust row numbers
                        for row in range(start_row, start_row + 14):  # Assume header is within first 14 rows
                            cell = ws.cell(row=row, column=1)  # Check first column for labels
                            if cell.value and str(cell.value).strip().lower() == label.lower():
                                # Found the label in the copied header
                                row_number = cell.row
                                
                                if merge_range:
                                    start_col, end_col = merge_range.split(":")
                                    start_col_idx = column_index_from_string(start_col)
                                    end_col_idx = column_index_from_string(end_col)
                                    cell_range = f"{start_col}{row_number}:{end_col}{row_number}"
                                    
                                    # First unmerge if already merged
                                    if cell_range in ws.merged_cells.ranges:
                                        ws.unmerge_cells(cell_range)
                                    
                                    ws.merge_cells(cell_range)
                                    ws.cell(row=row_number, column=start_col_idx).value = distributor_data.get(key, "")
                                    ws.cell(row=row_number, column=start_col_idx).alignment = Alignment(vertical='center', horizontal='left')
                                else:
                                    ws[f"{column}{row_number}"].value = distributor_data.get(key, "")
                                
                                break
                
                # Fill data rows based on matched distributor code
                distributor_data = df1[df1['Distributor code'] == code]
                grouped_data = distributor_data.groupby('Item Type')
                
                data_start_row = start_row + 14  # Adjust based on header size
                
                for category, items in grouped_data:
                    ws.cell(row=data_start_row, column=2).value = category
                    for _, row in items.iterrows():
                        ws.cell(row=data_start_row, column=3).value = row['Item Name']
                        ws.cell(row=data_start_row, column=4).value = row['Item QTY As Per book Stock']
                        ws.cell(row=data_start_row, column=10).value = row['Total Physical Stock']

                        a_val = float(ws.cell(row=data_start_row, column=4).value or 0)
                        b_val = float(ws.cell(row=data_start_row, column=5).value or 0)
                        c_val = float(ws.cell(row=data_start_row, column=6).value or 0)
                        d_val = float(ws.cell(row=data_start_row, column=7).value or 0)
                        f_val = float(ws.cell(row=data_start_row, column=9).value or 0)
                        g_val = float(ws.cell(row=data_start_row, column=10).value or 0)
                        h_val = float(ws.cell(row=data_start_row, column=11).value or 0)

                        e_result = a_val + b_val - c_val - d_val
                        ws.cell(row=data_start_row, column=8).value = e_result

                        i_result = f_val + g_val + h_val
                        ws.cell(row=data_start_row, column=13).value = i_result

                        j_result = e_result - i_result
                        ws.cell(row=data_start_row, column=14).value = abs(j_result)

                        ws.cell(row=data_start_row, column=8).number_format = '#,##0'
                        ws.cell(row=data_start_row, column=13).number_format = '#,##0'
                        ws.cell(row=data_start_row, column=14).number_format = '#,##0'

                        data_start_row += 1

        # Calculate the last row with data for this report
        last_row = data_start_row - 1
        
        # Add grand total row
        grand_total_row = last_row + 1
        ws.cell(row=grand_total_row, column=2).value = "Grand Total"
        
        # Calculate totals
        # For the first report, use the original sum_column_after_marker
        if start_row == 1:
            ws.cell(row=grand_total_row, column=4).value = sum_column_after_marker(None, ws, "D", "A", start_row=start_row)
            ws.cell(row=grand_total_row, column=8).value = sum_column_after_marker(None, ws, "H", "E = A + B - C- D", start_row=start_row)
            ws.cell(row=grand_total_row, column=10).value = sum_column_after_marker(None, ws, "J", "F", start_row=start_row)
            ws.cell(row=grand_total_row, column=13).value = sum_column_after_marker(None, ws, "M", "I =  F + G + H", start_row=start_row)
            ws.cell(row=grand_total_row, column=14).value = sum_column_after_marker(None, ws, "N", "J = E - I", start_row=start_row)
        else:
            # For subsequent reports, manually sum the columns within the range
            col_sums = {
                4: 0,  # Column D
                8: 0,  # Column H
                10: 0, # Column J
                13: 0, # Column M
                14: 0  # Column N
            }
            
            # Sum values from the data rows
            for row in range(start_row + 14, last_row + 1):
                for col in col_sums.keys():
                    val = ws.cell(row=row, column=col).value
                    if isinstance(val, (int, float)):
                        col_sums[col] += val
            
            # Write totals
            for col, total in col_sums.items():
                ws.cell(row=grand_total_row, column=col).value = total
                ws.cell(row=grand_total_row, column=col).number_format = '#,##0'
        
        return grand_total_row + 1  # Return the row after the grand total
    
    except Exception as e:
        print(f"Error processing report: {e}")
        return start_row


def fill_template_with_multiple_data(input_files):
    """
    Generate multiple reports in a single Excel file
    
    input_files: list of tuples [(df1_path1, df2_path1), (df1_path2, df2_path2), ...]
    """
    try:
        new_file = copy_template()
        if not new_file:
            return
        
        wb = load_workbook(new_file)
        ws = wb.active
        
        current_row = 1
        serial_number = 1
        
        for df1_path, df2_path in input_files:
            print(f"\nProcessing report #{serial_number} with files:")
            print(f"  - Data file: {df1_path}")
            print(f"  - Distributor file: {df2_path}")
            
            # Process one report starting at current_row
            end_row = process_single_report(ws, df1_path, df2_path, current_row, serial_number)
            
            # Move to next position with 3 blank rows
            current_row = end_row + 3
            serial_number += 1
        
        wb.save(new_file)
        print(f"\nAll reports generated successfully in: {new_file}")
        
    except Exception as e:
        print(f"Error processing files: {e}")


if __name__ == "__main__":
    # List of input file pairs to process
    input_files = [
        (r"Excel_1\ExampleFile (GUJ35).xlsx", r'Excel_2\dist_data.xlsx'),
        # Add more file pairs as needed
    ]
    
    fill_template_with_multiple_data(input_files)
