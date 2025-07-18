from openpyxl import load_workbook
import os

# Yeh dict: label text -> data key
labels_with_column = {
    "Report Number":        ("report_number", "D"),
    "Zone":                 ("zone", "D"),
    "Unit":                 ("unit", "D"),
    "Customer Code of the Distributor": ("customer_code", "D"),
    "Name of the Distributor": ("distributor_name", "D"),
    "Address 1":            ("address1", "D"),
    "Address 2":            ("address2", "L"),
    "City, State":           ("city_state", "L"),
    "Audit Firm":           ("audit_firm", "L"),
    "Audit Team Lead":      ("audit_team_lead", "L"),
    "Contact No":           ("contact_no", "L"),
    "Date of Audit":        ("date_of_audit", "L"),
}

def find_and_fill(ws, label, column_letter, value):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == label.lower():
                row_number = cell.row
                target_cell = ws[f"{column_letter}{row_number}"]
                target_cell.value = value
                return

def fill_header(ws, distributor_data):
    for label, (key, column) in labels_with_column.items():
        value = distributor_data.get(key, "")
        find_and_fill(ws, label, column, value)

# Load template
wb = load_workbook("template/template.xlsx")
ws = wb.active

# Distributor sample data
distributor = {
    "report_number": 1,
    "zone": "GMP",
    "unit": "",
    "customer_code": "500290430",
    "distributor_name": "SIDDHIVINAYAK SALES CORPORATION",
    "address1": "509/26 PARMANAND PATEL",
    "address2": "CHOWK NR ANJAR CINEMA",
    "city_state": "AHMADABAD, GUJARAT",
    "audit_firm": "RUTUL SHAH & ASSOCIATES",
    "audit_team_lead": "",
    "contact_no": "",
    "date_of_audit": "",
}

fill_header(ws, distributor)

# Save final file
wb.save("filled_template.xlsx")
print("✅ Saved filled_template.xlsx")
