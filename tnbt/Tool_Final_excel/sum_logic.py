from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def sum_column_after_marker(file_path, ws, column_letter, marker_value, start_row=None):
    col_idx = column_index_from_string(column_letter)
    start_sum = False
    sum_total = 0
    section_end = None

    # Find the next "HINDUSTAN COCA-COLA..." marker to determine section end
    for row in ws.iter_rows():
        if row[0].value and "HINDUSTAN COCA-COLA" in str(row[0].value):
            if start_sum:  # We found the next section's start
                section_end = row[0].row - 1
                break
    
    # If no next section found, use last row
    if not section_end:
        section_end = ws.max_row

    # Start from provided start_row or from beginning
    min_row = start_row if start_row else 1
    
    for row in ws.iter_rows(min_row=min_row, max_row=section_end, max_col=col_idx):
        cell_value = row[col_idx - 1].value

        if cell_value == marker_value:
            start_sum = True
            continue

        if start_sum:
            try:
                val = float(cell_value)
                sum_total += val
            except (TypeError, ValueError):
                continue

    return sum_total