import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Fill, Protection
import shutil
import os
from datetime import datetime
from openpyxl.utils import column_index_from_string
from openpyxl.cell.cell import MergedCell
from sum_logic import sum_column_after_marker

# Label text -> (data key, target column letter)
labels_with_column = {
    "Report Number":        ("report_number", "D", "D:G"),
    "Zone":                 ("zone", "D", "D:G"),
    "Unit":                 ("unit", "D", "D:G"),
    "Customer Code of the Distributor": ("customer_code", "D", "D:G"),
    "Name of the Distributor": ("distributor_name", "D", "D:G"),
    "Address 1":            ("address1", "D", "D:G"),
    "Address 2":            ("address2", "L", "L:O"),
    "City, State":          ("city_state", "L", "L:O"),
    "Audit Firm":           ("audit_firm", "L", "L:O"),
    "Audit Team Lead :":      ("audit_team_lead", "L", "L:O"),
    "Contact No":           ("contact_no", "L", "L:O"),
    "Date of audit":        ("date_of_audit", "L", "L:O")
}


def copy_template():
    template_path = os.path.join("template", "template.xlsx")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    new_filename = f"template_copy_{timestamp}.xlsx"
    
    if not os.path.exists(template_path):
        print(f"Template file not found at: {template_path}")
        return None
    
    try:
        shutil.copy2(template_path, new_filename)
        print(f"Excel file copied successfully to: {new_filename}")
        return new_filename
    except Exception as e:
        print(f"Error copying file: {e}")
        return None

def find_and_fill(ws, label, column_letter, value, merge_range=None):
    """Find row by label and fill value in fixed column and optionally merge."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == label.lower():
                row_number = cell.row

                if merge_range:
                    start_col, end_col = merge_range.split(":")
                    start_col_idx = column_index_from_string(start_col)
                    end_col_idx = column_index_from_string(end_col)
                    cell_range = f"{start_col}{row_number}:{end_col}{row_number}"

                    # First unmerge if already merged
                    if cell_range in ws.merged_cells.ranges:
                        ws.unmerge_cells(cell_range)

                    ws.merge_cells(cell_range)
                    ws.cell(row=row_number, column=start_col_idx).value = value
                    ws.cell(row=row_number, column=start_col_idx).alignment = Alignment(vertical='center', horizontal='left')
                else:
                    ws[f"{column_letter}{row_number}"].value = value

                return


def fill_header_section(ws, matched_row, serial_number):
    """Fill headers dynamically based on label names"""
    # Prepare values as per required keys
    customer_name = str(matched_row['Customer Name'])
    if '-' in customer_name:
        number_part, string_part = customer_name.split('-', 1)
    else:
        number_part = string_part = customer_name

    distributor_data = {
        "report_number": serial_number,
        "zone": matched_row['Zone'],
        "unit": "",
        "customer_code": number_part.strip(),
        "distributor_name": string_part.strip(),
        "address1": matched_row['Address 1'],
        "address2": matched_row['Address 2'],
        "city_state": f"{matched_row['District']}, {matched_row['State Name']}",
        "audit_firm": "RUTUL SHAH & ASSOCIATES",
        "audit_team_lead": matched_row.get('Audit Team Lead', ""),
        "contact_no": matched_row.get('Contact No', ""),
        "date_of_audit": matched_row.get('Date of Audit', "")
    }

    for label, (key, column, merge_range) in labels_with_column.items():
        find_and_fill(ws, label, column, distributor_data.get(key, ""), merge_range)


def find_data_start_row(ws):
    """Find the row where data entry should start by looking for specific header markers"""
    header_markers = ["S.No.", "Item Type", "Item Name", "A", "B", "C", "D"]
    
    for row_idx in range(1, 30):  # Search within a reasonable range
        for col_idx in range(1, 5):  # Check first few columns for headers
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value and any(marker.lower() == str(cell_value).strip().lower() for marker in header_markers):
                # Found header row, data starts on next row
                return row_idx + 1
    
    # Fallback to default if no markers found
    print("Warning: Could not find header markers. Using default start row (15).")
    return 15


def copy_cell_format(source_cell, target_cell):
    """Copy cell formatting safely without using direct StyleProxy assignment"""
    # Copy number format
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    
    # Copy alignment
    if source_cell.alignment:
        alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            textRotation=source_cell.alignment.textRotation,
            wrapText=source_cell.alignment.wrapText,
            shrinkToFit=source_cell.alignment.shrinkToFit,
            indent=source_cell.alignment.indent,
            relativeIndent=source_cell.alignment.relativeIndent,
            justifyLastLine=source_cell.alignment.justifyLastLine,
            readingOrder=source_cell.alignment.readingOrder
        )
        target_cell.alignment = alignment
    
    # Copy font attributes individually
    if source_cell.font:
        font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            underline=source_cell.font.underline,
            strike=source_cell.font.strike,
            color=source_cell.font.color
        )
        target_cell.font = font
    
    # Copy border if needed (simplified)
    if source_cell.border:
        target_cell.border = source_cell.border


def process_single_report(ws, df1_path, df2_path, report_start_row, serial_number):
    """Process a single report beginning at the given start row"""
    try:
        df1 = pd.read_excel(df1_path)
        df2 = pd.read_excel(df2_path)
        
        # Find start row for data (only for first report)
        base_start_row = find_data_start_row(ws) if report_start_row == 1 else None
        header_size = base_start_row - 1 if base_start_row else 14  # Default header size
        
        # For reports after the first one, clone the template
        if report_start_row > 1:
            # Clone header section (based on either detected or default header size)
            header_rows = header_size
            max_col = 20  # Reasonable default
            
            # Find maximum column used in header
            for row_idx in range(1, header_rows + 1):
                for col_idx in range(1, 30):
                    if ws.cell(row=row_idx, column=col_idx).value is not None:
                        max_col = max(max_col, col_idx)
            
            # Copy header section
            for row_idx in range(1, header_rows + 1):
                for col_idx in range(1, max_col + 1):
                    source_cell = ws.cell(row=row_idx, column=col_idx)
                    target_cell = ws.cell(row=report_start_row + row_idx - 1, column=col_idx)
                    
                    # Copy value
                    target_cell.value = source_cell.value
                    
                    # Copy formatting safely
                    copy_cell_format(source_cell, target_cell)
        
        for code in df1['Distributor code']:
            match = df2[df2['Codes'] == code]
            if not match.empty:
                matched_row = match.iloc[0]
                
                # Fill header section
                if report_start_row == 1:
                    fill_header_section(ws, matched_row, serial_number)
                else:
                    # For subsequent reports, adjust header position
                    customer_name = str(matched_row['Customer Name'])
                    if '-' in customer_name:
                        number_part, string_part = customer_name.split('-', 1)
                    else:
                        number_part = string_part = customer_name

                    distributor_data = {
                        "report_number": serial_number,
                        "zone": matched_row['Zone'],
                        "unit": "",
                        "customer_code": number_part.strip(),
                        "distributor_name": string_part.strip(),
                        "address1": matched_row['Address 1'],
                        "address2": matched_row['Address 2'],
                        "city_state": f"{matched_row['District']}, {matched_row['State Name']}",
                        "audit_firm": "RUTUL SHAH & ASSOCIATES",
                        "audit_team_lead": matched_row.get('Audit Team Lead', ""),
                        "contact_no": matched_row.get('Contact No', ""),
                        "date_of_audit": matched_row.get('Date of Audit', "")
                    }
                    
                    # Search for labels in the new header position
                    for label, (key, column, merge_range) in labels_with_column.items():
                        for row_idx in range(report_start_row, report_start_row + header_size):
                            cell_value = ws.cell(row=row_idx, column=1).value
                            if cell_value and str(cell_value).strip().lower() == label.lower():
                                row_number = row_idx
                                
                                if merge_range:
                                    start_col, end_col = merge_range.split(":")
                                    start_col_idx = column_index_from_string(start_col)
                                    end_col_idx = column_index_from_string(end_col)
                                    cell_range = f"{start_col}{row_number}:{end_col}{row_number}"
                                    
                                    # First unmerge if already merged
                                    if cell_range in ws.merged_cells.ranges:
                                        ws.unmerge_cells(cell_range)
                                    
                                    ws.merge_cells(cell_range)
                                    ws.cell(row=row_number, column=start_col_idx).value = distributor_data.get(key, "")
                                    ws.cell(row=row_number, column=start_col_idx).alignment = Alignment(vertical='center', horizontal='left')
                                else:
                                    ws[f"{column}{row_number}"].value = distributor_data.get(key, "")
                                
                                break
                
                # Calculate data start row
                if report_start_row == 1:
                    data_start_row = base_start_row
                else:
                    # For subsequent reports, use header offset + original data start row
                    data_start_row = report_start_row + header_size
                
                # Fill data rows
                distributor_data = df1[df1['Distributor code'] == code]
                grouped_data = distributor_data.groupby('Item Type')
                
                current_row = data_start_row
                
                for category, items in grouped_data:
                    ws.cell(row=current_row, column=2).value = category
                    for _, row in items.iterrows():
                        ws.cell(row=current_row, column=3).value = row['Item Name']
                        ws.cell(row=current_row, column=4).value = row['Item QTY As Per book Stock']
                        ws.cell(row=current_row, column=10).value = row['Total Physical Stock']

                        a_val = float(ws.cell(row=current_row, column=4).value or 0)
                        b_val = float(ws.cell(row=current_row, column=5).value or 0)
                        c_val = float(ws.cell(row=current_row, column=6).value or 0)
                        d_val = float(ws.cell(row=current_row, column=7).value or 0)
                        f_val = float(ws.cell(row=current_row, column=9).value or 0)
                        g_val = float(ws.cell(row=current_row, column=10).value or 0)
                        h_val = float(ws.cell(row=current_row, column=11).value or 0)

                        e_result = a_val + b_val - c_val - d_val
                        ws.cell(row=current_row, column=8).value = e_result

                        i_result = f_val + g_val + h_val
                        ws.cell(row=current_row, column=13).value = i_result

                        j_result = e_result - i_result
                        ws.cell(row=current_row, column=14).value = abs(j_result)

                        ws.cell(row=current_row, column=8).number_format = '#,##0'
                        ws.cell(row=current_row, column=13).number_format = '#,##0'
                        ws.cell(row=current_row, column=14).number_format = '#,##0'

                        current_row += 1
        
        # Last row with data
        last_row = current_row - 1
        
        # Add grand total row
        grand_total_row = last_row + 1
        ws.cell(row=grand_total_row, column=2).value = "Grand Total"
        
        # Calculate totals
        columns_to_sum = [4, 8, 10, 13, 14]  # D, H, J, M, N
        
        # Start row is the data start row
        start_sum_row = data_start_row
        
        for col in columns_to_sum:
            total = 0
            for row in range(start_sum_row, last_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if isinstance(cell_value, (int, float)):
                    total += cell_value
            
            ws.cell(row=grand_total_row, column=col).value = total
            ws.cell(row=grand_total_row, column=col).number_format = '#,##0'
        
        return grand_total_row + 1  # Return the row after the grand total
    
    except Exception as e:
        print(f"Error processing report: {e}")
        import traceback
        traceback.print_exc()
        return report_start_row


def fill_template_with_multiple_data(input_files):
    """
    Generate multiple reports in a single Excel file
    
    input_files: list of tuples [(df1_path1, df2_path1), (df1_path2, df2_path2), ...]
    """
    try:
        new_file = copy_template()
        if not new_file:
            return
        
        wb = load_workbook(new_file)
        ws = wb.active
        
        current_row = 1
        serial_number = 1
        
        for df1_path, df2_path in input_files:
            print(f"\nProcessing report #{serial_number} with files:")
            print(f"  - Data file: {df1_path}")
            print(f"  - Distributor file: {df2_path}")
            
            # Process one report starting at current_row
            end_row = process_single_report(ws, df1_path, df2_path, current_row, serial_number)
            
            # Move to next position with 3 blank rows
            current_row = end_row + 3
            serial_number += 1
        
        wb.save(new_file)
        print(f"\nAll reports generated successfully in: {new_file}")
        
    except Exception as e:
        print(f"Error processing files: {e}")


def fill_template_with_data():
    """Original function - now just calls the multiple data version with one set of files"""
    input_files = [
        (r"Excel_1\ExampleFile (GUJ35).xlsx", r'Excel_2\dist_data.xlsx')
    ]
    fill_template_with_multiple_data(input_files)


if __name__ == "__main__":
    # Option 1: Single report (original functionality)
    # fill_template_with_data()
    
    # Option 2: Multiple reports in one file
    input_files = [
        (r"Excel_1\ExampleFile (GUJ35).xlsx", r'Excel_2\dist_data.xlsx'),
        (r"Excel_1\ExampleFile (GUJ35).xlsx", r'Excel_2\dist_data.xlsx'),
        # Add more file pairs as needed
    ]
    fill_template_with_multiple_data(input_files)