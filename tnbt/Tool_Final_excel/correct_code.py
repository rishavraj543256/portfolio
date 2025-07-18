import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import shutil
import os
from datetime import datetime
from openpyxl.utils import column_index_from_string
from openpyxl.cell.cell import MergedCell
from sum_logic import sum_column_after_marker
from copy_temp import copy_template_to_existing_file

def select_file(title="Select File", filetypes=[("Excel files", "*.xlsx")]):
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=filetypes
    )
    return file_path

def get_output_preference():
    root = tk.Tk()
    root.withdraw()
    result = messagebox.askquestion(
        "Output Preference",
        "Do you want to generate a new output file?\n\n" +
        "Select 'Yes' for new file\n" +
        "Select 'No' to append to existing file"
    )
    return result == 'yes'

# Label text -> (data key, target column letter)
labels_with_column = {
    "Report Number":        ("report_number", "D", "D:G"),
    "Zone":                 ("zone", "D", "D:G"),
    "Unit":                 ("unit", "D", "D:G"),
    "Customer Code of the Distributor": ("customer_code", "D", "D:G"),
    "Name of the Distributor": ("distributor_name", "D", "D:G"),
    "Address 1":            ("address1", "D", "D:G"),
    "Address 2":            ("address2", "L", "L:O"),
    "City, State":          ("city_state", "L", "L:O"),
    "Audit Firm":           ("audit_firm", "L", "L:O"),
    "Audit Team Lead :":      ("audit_team_lead", "L", "L:O"),
    "Contact No":           ("contact_no", "L", "L:O"),
    "Date of audit":        ("date_of_audit", "L", "L:O")
}


def copy_template():
    template_path = os.path.join("template", "template.xlsx")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    new_filename = f"template_copy_{timestamp}.xlsx"
    
    if not os.path.exists(template_path):
        print(f"Template file not found at: {template_path}")
        return None
    
    try:
        shutil.copy2(template_path, new_filename)
        print(f"Excel file copied successfully to: {new_filename}")
        return new_filename
    except Exception as e:
        print(f"Error copying file: {e}")
        return None

def find_and_fill(ws, label, column_letter, value, merge_range=None):
    """Find row by label and fill value in fixed column and optionally merge."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == label.lower():
                row_number = cell.row

                if merge_range:
                    start_col, end_col = merge_range.split(":")
                    start_col_idx = column_index_from_string(start_col)
                    end_col_idx = column_index_from_string(end_col)
                    cell_range = f"{start_col}{row_number}:{end_col}{row_number}"

                    # First unmerge if already merged
                    if cell_range in ws.merged_cells.ranges:
                        #print(f"Unmerging existing range: {cell_range}")
                        ws.unmerge_cells(cell_range)

                    #print(f"Merging cells: {cell_range} with value: {value}")
                    ws.merge_cells(cell_range)
                    ws.cell(row=row_number, column=start_col_idx).value = value
                    ws.cell(row=row_number, column=start_col_idx).alignment = Alignment(vertical='center', horizontal='left')
                else:
                    ws[f"{column_letter}{row_number}"].value = value

                return



def fill_header_section(ws, matched_row, serial_number):
    """Fill headers dynamically based on label names"""
    # Prepare values as per required keys
    customer_name = str(matched_row['Customer Name'])
    if '-' in customer_name:
        number_part, string_part = customer_name.split('-', 1)
    else:
        number_part = string_part = customer_name

    distributor_data = {
        "report_number": serial_number,
        "zone": matched_row['Zone'],
        "unit": "",
        "customer_code": number_part.strip(),
        "distributor_name": string_part.strip(),
        "address1": matched_row['Address 1'],
        "address2": matched_row['Address 2'],
        "city_state": f"{matched_row['District']}, {matched_row['State Name']}",
        "audit_firm": "RUTUL SHAH & ASSOCIATES",
        "audit_team_lead": "",
        "contact_no": "",
        "date_of_audit": ""
    }

    for label, (key, column, merge_range) in labels_with_column.items():
        find_and_fill(ws, label, column, distributor_data.get(key, ""), merge_range)


def fill_template_with_data():
    try:
        # Select input files
        df1_path = select_file(title="Select ExampleFile Excel")
        if not df1_path:
            print("No ExampleFile selected. Exiting...")
            return
            
        df2_path = select_file(title="Select Distributor Data Excel")
        if not df2_path:
            print("No Distributor Data file selected. Exiting...")
            return

        # Load the input data
        df1 = pd.read_excel(df1_path)
        df2 = pd.read_excel(df2_path)

        # Get output preference
        create_new = get_output_preference()
        
        if create_new:
            new_file = copy_template()
            if not new_file:
                return
        else:
            existing_file = select_file(title="Select Existing Output File")
            if not existing_file:
                print("No output file selected. Exiting...")
                return
            copy_template_to_existing_file("template/template.xlsx", existing_file)
            new_file = existing_file

        wb = load_workbook(new_file)
        ws = wb.active
        
        serial_number = 1

        for code in df1['Distributor code']:
            match = df2[df2['Codes'] == code]
            if not match.empty:
                matched_row = match.iloc[0]
                fill_header_section(ws, matched_row, serial_number)
                #print("Header section filled for:", matched_row['Customer Name'])
                distributor_data = df1[df1['Distributor code'] == code]
                grouped_data = distributor_data.groupby('Item Type')
                start_row = 15
                
                for category, items in grouped_data:
                    ws.cell(row=start_row, column=2).value = category
                    for _, row in items.iterrows():
                        ws.cell(row=start_row, column=3).value = row['Item Name']
                        ws.cell(row=start_row, column=4).value = row['Item QTY As Per book Stock']
                        ws.cell(row=start_row, column=10).value = row['Total Physical Stock']

                        a_val = float(ws.cell(row=start_row, column=4).value or 0)
                        b_val = float(ws.cell(row=start_row, column=5).value or 0)
                        c_val = float(ws.cell(row=start_row, column=6).value or 0)
                        d_val = float(ws.cell(row=start_row, column=7).value or 0)
                        f_val = float(ws.cell(row=start_row, column=9).value or 0)
                        g_val = float(ws.cell(row=start_row, column=10).value or 0)
                        h_val = float(ws.cell(row=start_row, column=11).value or 0)

                        e_result = a_val + b_val - c_val - d_val
                        ws.cell(row=start_row, column=8).value = e_result

                        i_result = f_val + g_val + h_val
                        ws.cell(row=start_row, column=13).value = i_result

                        j_result = e_result - i_result
                        ws.cell(row=start_row, column=14).value = abs(j_result)

                        ws.cell(row=start_row, column=8).number_format = '#,##0'
                        ws.cell(row=start_row, column=13).number_format = '#,##0'
                        ws.cell(row=start_row, column=14).number_format = '#,##0'

                        start_row += 1

        last_row = 0
        for row in ws.iter_rows():
            if any(cell.value is not None for cell in row):
                last_row = row[0].row

        print(f"Last row with data: {last_row}")
        grand_total_row = last_row + 1
        ws.cell(row=grand_total_row, column=2).value = "Grand Total"
        
        sheet = wb.active
        ws.cell(row=grand_total_row, column=4).value = sum_column_after_marker(new_file, sheet, "D", "A")
        ws.cell(row=grand_total_row, column=8).value = sum_column_after_marker(new_file, sheet, "H", "E = A + B - C- D")
        ws.cell(row=grand_total_row, column=10).value = sum_column_after_marker(new_file, sheet, "J", "F")
        ws.cell(row=grand_total_row, column=13).value = sum_column_after_marker(new_file, sheet, "M", "I =  F + G + H")
        ws.cell(row=grand_total_row, column=14).value = sum_column_after_marker(new_file, sheet, "N", "J = E - I")

        wb.save(new_file)
        print(f"Template filled successfully: {new_file}")
        
    except Exception as e:
        print(f"Error processing files: {e}")

if __name__ == "__main__":
    fill_template_with_data()
