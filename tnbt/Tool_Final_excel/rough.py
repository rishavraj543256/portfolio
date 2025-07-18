import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import shutil
import os
from datetime import datetime
from openpyxl.utils import column_index_from_string
import tkinter as tk
from tkinter import filedialog, messagebox

def copy_template_format_to_existing(ws, start_row, template_path="template/template.xlsx"):
    """Copy the template format to the specified position in the worksheet"""
    try:
        template_wb = load_workbook(template_path)
        template_ws = template_wb.active
        
        # Copy merged cells
        for merged_range in template_ws.merged_cells.ranges:
            new_range = f"{merged_range.min_col_letter}{start_row + merged_range.min_row - 1}:{merged_range.max_col_letter}{start_row + merged_range.max_row - 1}"
            ws.merge_cells(new_range)
        
        # Copy row heights
        for row in range(1, template_ws.max_row + 1):
            if row in template_ws.row_dimensions:
                ws.row_dimensions[start_row + row - 1].height = template_ws.row_dimensions[row].height
        
        # Copy column widths
        for col in range(1, template_ws.max_column + 1):
            col_letter = get_column_letter(col)
            if col_letter in template_ws.column_dimensions:
                ws.column_dimensions[col_letter].width = template_ws.column_dimensions[col_letter].width
        
        # Copy cell styles for header section
        header_rows = 20  # Number of rows in the header section of template
        for row in range(1, header_rows + 1):
            for col in range(1, template_ws.max_column + 1):
                template_cell = template_ws.cell(row=row, column=col)
                new_cell = ws.cell(row=start_row + row - 1, column=col)
                
                if template_cell.has_style:
                    new_cell.font = Font(
                        name=template_cell.font.name,
                        size=template_cell.font.size,
                        bold=template_cell.font.bold,
                        italic=template_cell.font.italic
                    )
                    new_cell.alignment = Alignment(
                        horizontal=template_cell.alignment.horizontal,
                        vertical=template_cell.alignment.vertical,
                        wrap_text=template_cell.alignment.wrap_text
                    )
                    new_cell.border = template_cell.border
                    new_cell.fill = template_cell.fill
                    new_cell.number_format = template_cell.number_format
        
        return True
    except Exception as e:
        print(f"Error copying template format: {e}")
        return False

def select_excel_files():
    try:
        # [Previous code remains the same until the processing loop]
        
        # Process each Excel file in the folder
        for excel_file in excel_files:
            try:
                print(f"\nProcessing file: {excel_file}")
                df1 = pd.read_excel(excel_file)
                
                # Find all unique distributor codes in this file
                unique_codes = df1['Distributor code'].unique()
                print(f"Found distributor codes in file: {unique_codes}")
                
                # Track if we found any matching codes
                found_any_match = False
                
                for code in unique_codes:
                    # Find matching distributor in df2
                    match = df2[df2['Code Of DB'] == code]
                    
                    if not match.empty:
                        found_any_match = True
                        matched_row = match.iloc[0]
                        
                        print(f"Processing distributor code: {code}")
                        
                        # For existing files, copy the template format to the new section
                        if not create_new:
                            # First add 3 blank rows for spacing
                            ws.insert_rows(new_start_row, amount=3)
                            # Then copy the template format
                            if not copy_template_format_to_existing(ws, new_start_row):
                                print("Warning: Failed to copy template format")
                        
                        # [Rest of the processing code remains the same]
                        
            except Exception as e:
                print(f"Error processing file {excel_file}: {str(e)}")
                continue

        print(f"\nAll files processed successfully. Output saved to: {new_file}")
        
    except Exception as e:
        print(f"Error processing files: {str(e)}")























def process_excel_files():
    try:
        # [Previous code for file selection and setup remains the same...]
        
        # Process each Excel file in the folder
        for file_index, excel_file in enumerate(excel_files):
            try:
                print(f"\nProcessing file {file_index+1} of {len(excel_files)}: {excel_file}")
                df1 = pd.read_excel(excel_file)
                
                # Find all unique distributor codes in this file
                unique_codes = df1['Distributor code'].unique()
                print(f"Found distributor codes: {unique_codes}")
                
                # Track if we found any matching codes
                found_any_match = False
                
                for code in unique_codes:
                    match = df2[df2['Code Of DB'] == code]
                    
                    if not match.empty:
                        found_any_match = True
                        matched_row = match.iloc[0]
                        
                        print(f"Processing distributor code: {code}")
                        
                        # For all files after the first one OR when appending to existing file,
                        # we need to copy the template format
                        if file_index > 0 or not create_new:
                            # Add spacing between reports
                            ws.insert_rows(new_start_row, amount=3)
                            
                            # Copy template format for the new report section
                            if not copy_template_section(ws, new_start_row):
                                print("Warning: Template format not fully applied")
                        
                        # [Rest of your processing code remains the same...]
                        
                        # After processing, update positions for next report
                        current_report_number += 1
                        new_start_row = ws.max_row + 3  # Space for next report
                        
                if not found_any_match:
                    print(f"No matching distributor codes found in file: {excel_file}")
                else:
                    wb.save(new_file)
                    
            except Exception as e:
                print(f"Error processing file {excel_file}: {str(e)}")
                continue

        print(f"\nAll files processed successfully. Output saved to: {new_file}")
        
    except Exception as e:
        print(f"Error processing files: {str(e)}")

def copy_template_section(ws, start_row, template_path="template/template.xlsx"):
    """Copy a full report section from template to specified position"""
    try:
        template_wb = load_workbook(template_path)
        template_ws = template_wb.active
        
        # Determine how many rows to copy (entire report section)
        section_rows = 0
        for row in template_ws.iter_rows():
            if any(cell.value for cell in row):
                section_rows += 1
            else:
                break  # Stop at first empty row
        
        # Copy cell values and formatting
        for row_idx in range(1, section_rows + 1):
            # Insert new row at target position
            ws.insert_rows(start_row + row_idx - 1)
            
            # Copy each cell's value and formatting
            for col_idx in range(1, template_ws.max_column + 1):
                src_cell = template_ws.cell(row=row_idx, column=col_idx)
                tgt_cell = ws.cell(row=start_row + row_idx - 1, column=col_idx)
                
                # Copy value
                tgt_cell.value = src_cell.value
                
                # Copy formatting if cell has style
                if src_cell.has_style:
                    tgt_cell.font = Font(
                        name=src_cell.font.name,
                        size=src_cell.font.size,
                        bold=src_cell.font.bold,
                        italic=src_cell.font.italic
                    )
                    tgt_cell.alignment = Alignment(
                        horizontal=src_cell.alignment.horizontal,
                        vertical=src_cell.alignment.vertical,
                        wrap_text=src_cell.alignment.wrap_text
                    )
                    tgt_cell.border = src_cell.border
                    tgt_cell.fill = src_cell.fill
                    tgt_cell.number_format = src_cell.number_format
        
        # Copy merged cells for this section
        for merged_range in template_ws.merged_cells.ranges:
            if merged_range.min_row <= section_rows:
                new_range = (
                    f"{merged_range.min_col_letter}{start_row + merged_range.min_row - 1}:"
                    f"{merged_range.max_col_letter}{start_row + merged_range.max_row - 1}"
                )
                ws.merge_cells(new_range)
        
        # Copy row heights
        for row_idx in range(1, section_rows + 1):
            if row_idx in template_ws.row_dimensions:
                ws.row_dimensions[start_row + row_idx - 1].height = (
                    template_ws.row_dimensions[row_idx].height
                )
        
        return True
        
    except Exception as e:
        print(f"Error copying template section: {e}")
        return False






