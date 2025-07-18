from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

def copy_template_to_existing_file(template_path, existing_output_path):
    template_wb = load_workbook(template_path)
    template_ws = template_wb.active

    output_wb = load_workbook(existing_output_path)
    output_ws = output_wb.active

    # Step 1: Find the last used row
    last_row = output_ws.max_row
    new_start_row = last_row + 4  # 3 blank rows + 1 to start

    # Step 2: Copy row by row
    for row in template_ws.iter_rows():
        target_row_idx = new_start_row + row[0].row - 1
        for cell in row:
            new_cell = output_ws.cell(row=target_row_idx, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format

    # Step 3: Copy merged cells
    for merged_cell_range in template_ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_cell_range.bounds
        shifted_range = f"{get_column_letter(min_col)}{min_row + new_start_row - 1}:{get_column_letter(max_col)}{max_row + new_start_row - 1}"
        output_ws.merge_cells(shifted_range)

    # Step 4: Copy column widths
    for col in template_ws.column_dimensions:
        output_ws.column_dimensions[col].width = template_ws.column_dimensions[col].width

    # Step 5: Copy row heights
    for row_dim in template_ws.row_dimensions:
        output_ws.row_dimensions[new_start_row + row_dim - 1].height = template_ws.row_dimensions[row_dim].height

    output_wb.save(existing_output_path)
    print(f"✅ Template format copied to {existing_output_path} after row {last_row}")


    return new_start_row
#copy_template_to_existing_file("template/template.xlsx", "template_copy_20250407_220135.xlsx")